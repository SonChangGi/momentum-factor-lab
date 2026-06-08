import json
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from momentum_factor_lab.config import RunConfig
from momentum_factor_lab.data import MarketData, generate_offline_sample_data
from momentum_factor_lab.report import _executive_summary_lines, write_pdf, write_reports
from momentum_factor_lab.workflow import run_analysis, write_run_results_json


LIQUIDITY_CAPACITY_COLUMNS = {
    "avg_share_volume_63d",
    "avg_dollar_volume_63d",
    "price_observations_63d",
    "volume_observations_63d",
    "dollar_volume_observations_63d",
    "min_liquidity_observations_required",
    "min_avg_volume_required",
    "min_avg_dollar_volume_required",
    "liquidity_evidence_status",
    "liquidity_filter_pass",
    "data_quality_status",
    "data_quality_pass",
    "data_quality_warning",
    "proposed_weight",
    "target_aum",
    "max_adv_participation",
    "target_notional",
    "max_trade_notional_by_adv",
    "capacity_notional_limit",
    "capacity_aum_limit",
    "adv_participation",
    "capacity_utilization",
    "capacity_status",
    "capacity_pass",
    "capacity_warning",
}

WEIGHTING_COLUMNS = {
    "weighting_method",
    "raw_weight_score",
    "pre_cap_weight",
    "weight_cap",
    "weight_cap_excess",
    "score_component",
    "market_cap",
    "market_cap_source",
    "market_cap_component",
    "liquidity_component",
    "size_component_source",
}


def _live_fixture_market_data(
    config: RunConfig,
    *,
    subset_run: bool,
    point_in_time: bool = False,
    requested_symbols: int | None = None,
    eligible_symbols: int | None = None,
) -> MarketData:
    sample_config = replace(
        config,
        offline_sample=True,
        end_date=datetime.now(UTC).date().isoformat(),
    )
    sample = generate_offline_sample_data(sample_config)
    candidate_symbols = len(sample.candidate_universe)
    eligible_stock_symbols = len(sample.eligible_universe)
    if requested_symbols is None:
        candidate_cap = max((config.max_price_symbols or (eligible_stock_symbols + 1)) - 1, 0)
        requested_symbols = min(eligible_stock_symbols, candidate_cap) if subset_run else eligible_stock_symbols
    if eligible_symbols is None:
        eligible_symbols = eligible_stock_symbols
    summary = pd.DataFrame(
        [
            {
                "source": "live-run-summary",
                "status": "partial_subset" if subset_run else "full_requested_universe",
                "records": len(sample.prices.columns),
                "candidate_symbols": candidate_symbols,
                "requested_price_symbols": requested_symbols,
                "eligible_price_symbols": eligible_symbols,
                "requested_download_symbols": requested_symbols + 1,
                "benchmark_symbol": config.benchmark,
                "benchmark_price_available": config.benchmark in sample.prices.columns,
                "excluded_symbols": 0,
                "subset_run": subset_run,
                "point_in_time_universe": point_in_time,
            }
        ]
    )
    if point_in_time:
        summary = pd.concat(
            [
                summary,
                pd.DataFrame(
                    [
                        {
                            "source": "user-point-in-time-universe-provenance",
                            "status": "attested",
                            "records": candidate_symbols,
                            "candidate_symbols": candidate_symbols,
                            "point_in_time_universe": True,
                            "tradable_universe_approved": bool(config.approved_tradable_universe),
                            "universe_provenance": "source=test-fixture as_of=2026-06-05 symbol_count=2456 hash=fixture",
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )
    return MarketData(
        prices=sample.prices,
        volumes=sample.volumes,
        provider="test-live-data",
        fetched_at=datetime.now(UTC),
        as_of=sample.prices.index.max(),
        exclusions=sample.exclusions,
        offline_sample=False,
        candidate_universe=sample.candidate_universe,
        eligible_universe=sample.eligible_universe,
        price_sources=sample.price_sources,
        data_sources=summary,
    )


def test_offline_run_generates_required_outputs(tmp_path):
    config = RunConfig(
        start_date="2019-01-01",
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=True,
    )
    result = write_reports(run_analysis(config))
    pdf = Path(result.output_paths["pdf"])
    xlsx = Path(result.output_paths["xlsx"])
    json_path = Path(result.output_paths["json"])
    assert pdf.exists() and pdf.stat().st_size > 5_000
    assert xlsx.exists() and xlsx.stat().st_size > 5_000
    assert json_path.exists()
    workbook = openpyxl.load_workbook(xlsx, read_only=True)
    required = {
        "config_assumptions",
        "universe",
        "factor_scores",
        "factor_score_history_top20",
        "data_sources",
        "price_sources",
        "data_quality",
        "factor_definitions",
        "factor_validation",
        "backtest_metrics",
        "score_components",
        "benchmark_relative",
        "selected_factor_scores",
        "selected_factor",
        "recommendations",
        "exclusions",
        "robustness",
        "sensitivity",
        "status_panel",
        "factor_family_leaderboard",
        "factor_overlap_top20",
        "regime_performance",
        "tradability_gate",
        "universe_provenance",
        "liquidity_capacity",
        "cost_stress",
        "selection_history",
    }
    assert required.issubset(set(workbook.sheetnames))
    assert result.metadata["recommendation_status"] == "sample_offline_not_current"
    assert not result.metadata["current_recommendations_available"]
    assert "validation" in set(result.robustness["slice"])
    assert result.score_components.index[0] == result.selected_factor
    assert not result.benchmark_relative.empty
    assert len(result.recommendations) == result.config.top_n
    assert result.metadata["candidate_universe_size"] >= 2000
    assert result.metadata["eligible_price_universe_size"] == len(result.market_data.eligible_universe)
    assert result.metadata["benchmark_symbol"] == "SPY"
    assert result.metadata["benchmark_price_available"]
    assert "SPY" in result.market_data.prices.columns
    assert "SPY" not in result.market_data.eligible_universe["symbol"].to_list()
    assert not result.market_data.candidate_universe["is_etf"].any()
    assert {"SPY", "QQQ"}.isdisjoint(set(result.recommendations["symbol"]))
    for scores in result.factor_scores.values():
        assert {"SPY", "QQQ"}.isdisjoint(set(scores.columns))
    for backtest in result.backtests.values():
        assert {"SPY", "QQQ"}.isdisjoint(set(backtest.weights.columns))
    assert not result.benchmark_relative.empty
    assert set(result.benchmark_relative["benchmark"]) == {"SPY"}
    assert result.metadata["factor_count"] >= 55
    assert result.metadata["factor_validation_status"] == "pass"
    assert result.metadata["data_quality_manifest_available"]
    assert result.metadata["data_quality_status_counts"]
    assert "data_quality" in json.loads(json_path.read_text(encoding="utf-8"))
    assert {"variant_type", "variant", "parameter_set"}.issubset(result.sensitivity.columns)
    assert result.metadata["selected_factor_selection_source"] == "research_validation"
    assert not result.metadata["same_sample_selection_blocked_for_tradable"]
    assert WEIGHTING_COLUMNS.issubset(result.recommendations.columns)
    factor_sheet = workbook["factor_scores"]
    history_sheet = workbook["factor_score_history_top20"]
    assert factor_sheet.max_row > len(result.market_data.eligible_universe)
    assert history_sheet.max_row < 1_048_576


def test_raw_benchmark_and_nonbenchmark_etf_prices_do_not_enter_stock_analysis(tmp_path, monkeypatch):
    config = RunConfig(
        start_date="2019-01-01",
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=True,
    )
    market_data = generate_offline_sample_data(config)
    market_data.prices["QQQ"] = market_data.prices["SPY"] * 1.01
    market_data.volumes["QQQ"] = market_data.volumes["SPY"]
    market_data.price_sources = pd.concat(
        [market_data.price_sources, pd.DataFrame([{"symbol": "QQQ", "price_source": "fixture-etf-price"}])],
        ignore_index=True,
    )
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert {"SPY", "QQQ"}.issubset(set(result.market_data.prices.columns))
    assert {"SPY", "QQQ"}.isdisjoint(set(result.recommendations["symbol"]))
    assert {"SPY", "QQQ"}.isdisjoint(set(result.market_data.eligible_universe["symbol"]))
    for scores in result.factor_scores.values():
        assert {"SPY", "QQQ"}.isdisjoint(set(scores.columns))
    for backtest in result.backtests.values():
        assert {"SPY", "QQQ"}.isdisjoint(set(backtest.weights.columns))
    assert not result.benchmark_relative.empty
    assert set(result.benchmark_relative["benchmark"]) == {"SPY"}
    assert result.metadata["benchmark_price_available"]


def test_lowercase_benchmark_still_generates_benchmark_relative_metrics(tmp_path):
    result = run_analysis(
        RunConfig(
            start_date="2019-01-01",
            output_dir=tmp_path / "outputs",
            report_dir=tmp_path / "reports",
            offline_sample=True,
            benchmark="spy",
        )
    )

    assert result.metadata["benchmark_symbol"] == "SPY"
    assert result.metadata["benchmark_price_available"]
    assert not result.benchmark_relative.empty
    assert set(result.benchmark_relative["benchmark"]) == {"SPY"}


def test_missing_eligible_universe_fails_closed_before_analysis(tmp_path, monkeypatch):
    config = RunConfig(
        start_date="2019-01-01",
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    market_data.eligible_universe = pd.DataFrame(columns=["not_symbol"])
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    with pytest.raises(ValueError, match="no stock-only analysis price symbols"):
        run_analysis(config)


def test_live_fixture_counts_candidate_coverage_without_benchmark(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    summary = result.data_sources[result.data_sources["source"].eq("live-run-summary")].iloc[-1]
    assert int(summary["eligible_price_symbols"]) == len(result.market_data.eligible_universe)
    assert int(summary["requested_price_symbols"]) == len(result.market_data.eligible_universe)
    assert int(summary["records"]) == len(result.market_data.prices.columns)
    assert result.metadata["eligible_price_universe_size"] == len(result.market_data.eligible_universe)
    assert result.metadata["fetched_price_symbol_count"] == len(result.market_data.prices.columns)
    assert result.metadata["benchmark_price_available"]


def test_live_failure_or_offline_status_never_claims_current(tmp_path, monkeypatch):
    config = RunConfig(output_dir=tmp_path / "outputs", report_dir=tmp_path / "reports", offline_sample=True)
    result = run_analysis(config)
    assert result.metadata["recommendation_status"] != "current_live"
    assert not result.metadata["current_recommendations_available"]


def test_subset_live_run_exports_recommendations_with_subset_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        max_price_symbols=8,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=True),
    )

    result = write_reports(run_analysis(config))

    assert "subset_run" in result.metadata["recommendation_status"]
    assert result.metadata["current_recommendations_available"]
    assert result.metadata["tradable_recommendations_available"]
    assert result.metadata["recommendation_output_key"] == "recommendations"
    assert not result.metadata["research_only"]
    assert "no_explicit_price_symbol_cap" in result.metadata["execution_limitations"]
    assert result.recommendations["weight"].sum() > 0.0
    assert set(result.recommendations["recommendation_output"]) == {"recommendations"}
    workbook = openpyxl.load_workbook(result.output_paths["xlsx"], read_only=True)
    assert "recommendations" in workbook.sheetnames
    assert "research_signals" not in workbook.sheetnames


def test_full_live_run_without_point_in_time_evidence_exports_recommendations_with_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=False),
    )

    result = run_analysis(config)

    assert "with_limitations" in result.metadata["recommendation_status"]
    assert result.metadata["current_recommendations_available"]
    assert "point_in_time_universe" in result.metadata["execution_limitations"]
    assert result.metadata["recommendation_output_key"] == "recommendations"
    assert result.recommendations["weight"].sum() > 0.0


def test_current_live_source_flag_never_satisfies_point_in_time_gate(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=False)
    market_data.data_sources.loc[
        market_data.data_sources["source"].eq("live-run-summary"),
        ["point_in_time_universe", "universe_provenance"],
    ] = [True, "current live source/profile as-of run date"]
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert result.metadata["tradability_requirements"]["point_in_time_universe"] is False
    assert "point_in_time_universe" in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_unstructured_pit_attestation_does_not_satisfy_practical_gate(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    market_data.data_sources.loc[
        market_data.data_sources["source"].eq("user-point-in-time-universe-provenance"),
        "universe_provenance",
    ] = "test PIT membership provenance as-of 2026-06-05"
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert result.metadata["tradability_requirements"]["point_in_time_universe"] is False
    assert "point_in_time_universe" in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_fresh_live_run_uses_validation_selected_factor_without_false_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert result.metadata["selected_factor_selection_source"] == "research_validation"
    assert result.selected_factor == result.metadata["validation_selected_factor"]
    assert result.metadata["current_recommendations_available"]
    assert result.metadata["tradable_recommendations_available"]
    assert not result.metadata["same_sample_selection_blocked_for_tradable"]
    assert "predeclared_selected_factor" not in result.metadata["execution_limitations"]
    assert all("same_sample" not in item for item in result.metadata["execution_limitations"])
    assert "same_sample" not in result.metadata["recommendation_status"]
    assert result.recommendations["weight"].gt(0.0).any()
    assert WEIGHTING_COLUMNS.issubset(result.recommendations.columns)
    assert result.recommendations["weighting_method"].eq("score_size_liquidity").all()


def test_validation_ranked_factor_exports_recommendations_without_same_sample_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = write_reports(run_analysis(config))

    assert result.metadata["selected_factor_selection_source"] == "research_validation"
    assert result.metadata["recommendation_output_key"] == "recommendations"
    assert result.metadata["tradable_recommendations_available"]
    assert result.metadata["execution_limitations"] == []
    assert result.metadata["tradability_blockers"] == []
    assert not result.metadata["same_sample_selection_blocked_for_tradable"]
    assert result.recommendations["weight"].gt(0.0).any()
    assert result.recommendations["weight"].nunique() > 1
    assert WEIGHTING_COLUMNS.issubset(result.recommendations.columns)

    payload = json.loads(Path(result.output_paths["json"]).read_text(encoding="utf-8"))
    assert "recommendations" in payload
    assert "research_signals" not in payload
    assert {row["selected_factor_selection_source"] for row in payload["recommendations"]} == {"research_validation"}
    assert any(row["weight"] > 0 for row in payload["recommendations"])
    assert "predeclared_selected_factor" not in payload["metadata"]["execution_limitations"]
    assert WEIGHTING_COLUMNS.issubset(payload["recommendations"][0])

    workbook = openpyxl.load_workbook(Path(result.output_paths["xlsx"]), read_only=True)
    assert "recommendations" in workbook.sheetnames
    assert "research_signals" not in workbook.sheetnames
    headers = [cell.value for cell in next(workbook["recommendations"].iter_rows(max_row=1))]
    assert WEIGHTING_COLUMNS.issubset(headers)


def test_market_cap_enrichment_does_not_change_price_universe_or_coverage(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    candidate_count = len(market_data.candidate_universe)
    eligible_count = len(market_data.eligible_universe)
    price_columns = set(market_data.prices.columns)
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert len(result.market_data.candidate_universe) == candidate_count
    assert len(result.market_data.eligible_universe) == eligible_count
    assert set(result.market_data.prices.columns) == price_columns
    summary = result.data_sources[result.data_sources["source"].eq("live-run-summary")].iloc[-1]
    assert int(summary["requested_price_symbols"]) == eligible_count
    assert int(summary["eligible_price_symbols"]) == eligible_count
    assert "yfinance-fast-info-market-cap" in set(result.data_sources["source"])
    assert result.metadata["current_recommendations_available"]


def test_predeclared_factor_controls_fresh_live_recommendations_not_validation_rank(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )

    def rank_other_factor_first(metrics):
        ranked = metrics.assign(composite_score=0.0)[["composite_score"]]
        ranked.loc["mom_12_1", "composite_score"] = 1.0
        return ranked.sort_values("composite_score", ascending=False)

    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )
    monkeypatch.setattr("momentum_factor_lab.workflow._score_factors", rank_other_factor_first)

    result = run_analysis(config)

    assert result.metadata["validation_selected_factor"] == "mom_12_1"
    assert result.selected_factor == "mom_1m"
    assert result.metadata["factor_selection_mode"] == "predeclared"
    assert result.metadata["selected_factor_selection_source"] == "predeclared"
    assert not result.metadata["same_sample_selection_blocked_for_tradable"]
    assert result.metadata["recommendation_status"] == "current_live"
    assert result.metadata["current_recommendations_available"]
    assert result.metadata["tradable_recommendations_available"]
    assert result.metadata["tradability_blockers"] == []
    assert result.recommendations["selected_factor"].eq("mom_1m").all()
    assert result.recommendations["selected_factor_selection_source"].eq("predeclared").all()
    assert LIQUIDITY_CAPACITY_COLUMNS.issubset(result.recommendations.columns)
    assert result.recommendations["avg_share_volume_63d"].notna().all()
    assert result.recommendations["avg_dollar_volume_63d"].notna().all()
    assert result.recommendations["liquidity_evidence_status"].eq("pass").all()
    assert result.recommendations["data_quality_status"].eq("pass").all()
    assert result.recommendations["data_quality_pass"].all()
    assert result.recommendations["capacity_status"].eq("pass").all()
    assert result.recommendations["target_notional"].gt(0).all()


def test_walk_forward_factor_selection_records_limitation_without_frozen_policy(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        factor_selection_mode="walk_forward",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert result.metadata["selected_factor_selection_source"] == "walk_forward"
    assert not result.metadata["selection_policy_frozen_for_live"]
    assert "predeclared_selected_factor" not in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_current_live_predeclared_exports_recommendations_key_and_sheet(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = write_reports(run_analysis(config))
    for kind, output_path in result.output_paths.items():
        path = Path(output_path)
        assert path.exists()
        assert path.parent in {config.output_dir, config.report_dir}
        if kind == "json":
            assert path.name.startswith("run_results_")
            assert path.suffix == ".json"
        elif kind == "pdf":
            assert path.name.startswith("momentum_factor_report_")
            assert path.suffix == ".pdf"
        elif kind == "xlsx":
            assert path.name.startswith("momentum_factor_analysis_")
            assert path.suffix == ".xlsx"

    payload = json.loads(Path(result.output_paths["json"]).read_text(encoding="utf-8"))
    workbook = openpyxl.load_workbook(result.output_paths["xlsx"], read_only=True)

    assert result.metadata["recommendation_output_key"] == "recommendations"
    assert payload["metadata"]["recommendation_output_key"] == "recommendations"
    assert "recommendations" in payload
    assert "research_signals" not in payload
    assert "data_quality" in payload
    assert "recommendations" in workbook.sheetnames
    assert "research_signals" not in workbook.sheetnames
    assert "data_quality" in workbook.sheetnames


def test_stale_live_data_zeroes_recommendation_weights(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=1,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    market_data.as_of = pd.Timestamp("2020-01-01")
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert result.metadata["recommendation_status"].startswith("stale_live_data_")
    assert not result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].eq(0.0).all()
    assert result.recommendations["recommendation_status"].eq(result.metadata["recommendation_status"]).all()
    assert result.recommendations["signal_date"].eq(str(market_data.prices.index.max().date())).all()


def test_missing_volume_data_has_explicit_liquidity_capacity_warning(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        min_avg_volume=0,
        min_avg_dollar_volume=0,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    market_data.volumes = pd.DataFrame(index=market_data.prices.index)
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert LIQUIDITY_CAPACITY_COLUMNS.issubset(result.recommendations.columns)
    assert result.recommendations["liquidity_evidence_status"].eq("missing_liquidity_evidence").all()
    assert not result.recommendations["liquidity_filter_pass"].any()
    assert result.recommendations["capacity_status"].eq("missing_or_failed_liquidity_evidence").all()
    assert result.metadata["tradability_blockers"] == []
    assert "row_level_liquidity_pass" in result.metadata["execution_limitations"]
    assert "capacity_estimated_and_pass" in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_data_quality_anomaly_blocks_live_tradable_output(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    stock_columns = [symbol for symbol in market_data.prices.columns if symbol in set(market_data.eligible_universe["symbol"])]
    anomaly_date = market_data.prices.index[-2]
    prior_date = market_data.prices.index[-3]
    market_data.prices.loc[anomaly_date, stock_columns] = market_data.prices.loc[prior_date, stock_columns] * 2.5
    market_data.data_quality = pd.DataFrame()
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert "row_level_data_quality_pass" in result.metadata["tradability_blockers"]
    assert not result.metadata["current_recommendations_available"]
    assert not result.recommendations["data_quality_pass"].all()
    assert result.recommendations["data_quality_status"].eq("extreme_return_anomaly").any()
    assert result.recommendations["weight"].eq(0.0).all()


def test_sparse_liquidity_observations_record_advisory_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    market_data.volumes.iloc[:-1, :] = pd.NA
    market_data.volumes.iloc[-1, :] = 1_000_000_000
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert result.recommendations["volume_observations_63d"].eq(1).all()
    assert result.recommendations["liquidity_evidence_status"].eq("insufficient_liquidity_observations").all()
    assert result.metadata["tradability_blockers"] == []
    assert "row_level_liquidity_pass" in result.metadata["execution_limitations"]
    assert "capacity_estimated_and_pass" in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_live_recommendations_record_missing_capacity_inputs_as_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert result.metadata["current_recommendations_available"]
    assert "capacity_estimated_and_pass" in result.metadata["execution_limitations"]
    assert result.recommendations["capacity_status"].eq("not_estimated_missing_aum_and_participation_limit").all()
    assert result.recommendations["weight"].gt(0.0).any()


def test_capacity_adv_participation_limit_records_advisory_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=10_000_000_000,
        max_adv_participation=0.0001,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert result.recommendations["liquidity_evidence_status"].eq("pass").all()
    assert result.recommendations["capacity_status"].eq("exceeds_adv_participation_limit").all()
    assert "capacity_estimated_and_pass" in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_small_user_universe_records_limitations_unless_approved_with_provenance(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
        universe=["SPY", "QQQ", "AAPL", "MSFT", "NVDA", "AMZN"],
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert len(result.market_data.candidate_universe) < result.config.min_tradable_universe_size
    assert result.metadata["current_recommendations_available"]
    assert "broad_or_approved_tradable_universe" in result.metadata["execution_limitations"]


def test_current_source_approval_cannot_label_small_universe_tradable(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
        universe=["AAPL", "MSFT", "NVDA", "AMZN", "META", "GOOGL"],
        approved_tradable_universe=True,
    )
    market_data = _live_fixture_market_data(config, subset_run=False, point_in_time=True)
    market_data.data_sources.loc[
        market_data.data_sources["source"].eq("user-point-in-time-universe-provenance"),
        "tradable_universe_approved",
    ] = False
    market_data.data_sources.loc[
        market_data.data_sources["source"].eq("live-run-summary"),
        ["tradable_universe_approved", "universe_provenance"],
    ] = [True, "current source/profile approval is not user universe provenance"]
    monkeypatch.setattr("momentum_factor_lab.workflow.load_market_data", lambda _: market_data)

    result = run_analysis(config)

    assert result.metadata["tradability_requirements"]["point_in_time_universe"] is True
    assert result.metadata["tradability_requirements"]["broad_or_approved_tradable_universe"] is False
    assert "broad_or_approved_tradable_universe" in result.metadata["execution_limitations"]
    assert result.metadata["recommendation_output_label"] == "Practical recommendations"
    assert result.metadata["recommendation_output_key"] == "recommendations"
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_small_user_universe_with_approved_provenance_can_use_tradable_labels(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
        universe=["AAPL", "MSFT", "NVDA", "AMZN", "META", "GOOGL"],
        approved_tradable_universe=True,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert len(result.market_data.candidate_universe) < result.config.min_tradable_universe_size
    assert result.metadata["tradability_blockers"] == []
    assert result.metadata["recommendation_output_label"] == "Practical recommendations"
    assert result.metadata["recommendation_output_key"] == "recommendations"
    assert result.metadata["current_recommendations_available"]
    assert result.metadata["tradable_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_incomplete_requested_price_coverage_records_subset_limitation(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(
            config,
            subset_run=False,
            point_in_time=True,
            requested_symbols=100,
            eligible_symbols=24,
        ),
    )

    result = run_analysis(config)

    assert "complete_requested_price_coverage" in result.metadata["execution_limitations"]
    assert "no_explicit_price_symbol_cap" not in result.metadata["execution_limitations"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_json_export_preserves_recommendation_status_and_source_contract(tmp_path):
    config = RunConfig(
        start_date="2019-01-01",
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=True,
    )
    result = run_analysis(config)
    json_path = tmp_path / "run_results.json"
    write_run_results_json(result, json_path)

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    output_key = payload["metadata"]["recommendation_output_key"]
    rows = payload[output_key]
    recommendation_status = payload["metadata"]["recommendation_status"]

    assert payload["metadata"]["current_recommendations_available"] is False
    assert output_key == "recommendations"
    assert rows
    assert LIQUIDITY_CAPACITY_COLUMNS.issubset(rows[0])
    assert payload["data_quality"]
    assert payload["metadata"]["data_quality_gate"]["manifest_available"] is True
    assert WEIGHTING_COLUMNS.issubset(rows[0])
    assert {row["recommendation_status"] for row in rows} == {recommendation_status}
    assert {row["signal_date"] for row in rows} == {payload["metadata"]["signal_date"]}
    assert {row["capacity_status"] for row in rows} == {"not_estimated_missing_aum_and_participation_limit"}
    assert payload["data_sources"]
    assert payload["price_sources"]


def test_json_export_is_strict_and_replaces_non_finite_values(tmp_path):
    result = run_analysis(
        RunConfig(
            start_date="2019-01-01",
            output_dir=tmp_path / "outputs",
            report_dir=tmp_path / "reports",
            offline_sample=True,
        )
    )
    result.metadata["nan_probe"] = float("nan")
    result.data_quality.loc[result.data_quality.index[0], "latest_price"] = float("inf")
    json_path = tmp_path / "run_results.json"

    write_run_results_json(result, json_path)

    text = json_path.read_text(encoding="utf-8")
    assert "NaN" not in text
    assert "Infinity" not in text
    payload = json.loads(text)
    assert payload["metadata"]["nan_probe"] is None
    assert payload["data_quality"][0]["latest_price"] is None


def test_json_export_preserves_selected_turnover_cost_metadata(tmp_path):
    config = RunConfig(
        start_date="2019-01-01",
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=True,
    )
    result = run_analysis(config)
    json_path = tmp_path / "run_results.json"
    write_run_results_json(result, json_path)

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    metadata = payload["metadata"]
    selected_metrics = result.metrics.loc[result.selected_factor]

    assert metadata["selected_factor_avg_turnover"] == selected_metrics["full_avg_turnover"]
    assert metadata["selected_factor_total_turnover"] == selected_metrics["full_total_turnover"]
    assert metadata["selected_factor_annualized_turnover"] == selected_metrics["full_annualized_turnover"]
    assert metadata["selected_factor_total_cost"] == selected_metrics["full_total_cost"]
    assert metadata["selected_factor_annualized_cost_drag"] == selected_metrics["full_annualized_cost_drag"]


def test_non_current_pdf_labels_output_as_reference_recommendations(tmp_path, monkeypatch):
    result = run_analysis(
        RunConfig(
            start_date="2019-01-01",
            output_dir=tmp_path / "outputs",
            report_dir=tmp_path / "reports",
            offline_sample=True,
        )
    )
    text_pages: list[tuple[str, list[str]]] = []
    table_titles: list[str] = []

    def capture_text_page(pdf, title, lines):
        text_pages.append((title, lines))

    def capture_table_page(pdf, title, frame, max_rows=18):
        table_titles.append(title)

    monkeypatch.setattr("momentum_factor_lab.report._text_page", capture_text_page)
    monkeypatch.setattr("momentum_factor_lab.report._table_page", capture_table_page)

    write_pdf(result, tmp_path / "report.pdf")

    executive_lines = "\n".join(text_pages[0][1])
    assert "Output type: Reference recommendations (not current)" in executive_lines
    assert "Execution limitations:" in executive_lines
    assert "Recommendation weighting: score_size_liquidity" in executive_lines
    assert "cash remainder:" in executive_lines
    assert "Liquidity/capacity:" in executive_lines
    assert "Reference recommendations (not current)" in table_titles


def test_pdf_summary_lines_include_weighting_method_and_cash_remainder(tmp_path):
    result = run_analysis(
        RunConfig(
            start_date="2019-01-01",
            output_dir=tmp_path / "outputs",
            report_dir=tmp_path / "reports",
            offline_sample=True,
        )
    )

    lines = "\n".join(_executive_summary_lines(result))

    assert "Recommendation weighting: score_size_liquidity" in lines
    assert "cash remainder:" in lines


def test_current_live_pdf_labels_tradable_output_and_diagnostics(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        selected_factor="mom_1m",
        target_aum=100_000,
        max_adv_participation=0.05,
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )
    result = run_analysis(config)
    text_pages: list[tuple[str, list[str]]] = []
    table_pages: dict[str, pd.DataFrame] = {}

    def capture_text_page(pdf, title, lines):
        text_pages.append((title, lines))

    def capture_table_page(pdf, title, frame, max_rows=18):
        table_pages[title] = frame

    monkeypatch.setattr("momentum_factor_lab.report._text_page", capture_text_page)
    monkeypatch.setattr("momentum_factor_lab.report._table_page", capture_table_page)

    write_pdf(result, tmp_path / "report.pdf")

    executive_lines = "\n".join(text_pages[0][1])
    output_title = "Practical recommendations"
    assert f"Output type: {output_title}" in executive_lines
    assert "Execution limitations: none" in executive_lines
    assert "Recommendation weighting: score_size_liquidity" in executive_lines
    assert "cash remainder:" in executive_lines
    assert output_title in table_pages
    assert "Reference recommendations (not current)" not in table_pages
    assert {
        "recommendation_status",
        "signal_date",
        "recommendation_output",
        "selected_factor_selection_source",
    }.issubset(table_pages[output_title].columns)
    assert LIQUIDITY_CAPACITY_COLUMNS.issubset(table_pages[output_title].columns)


def test_report_exports_turnover_cost_diagnostic_columns(tmp_path):
    result = write_reports(
        run_analysis(
            RunConfig(
                start_date="2019-01-01",
                output_dir=tmp_path / "outputs",
                report_dir=tmp_path / "reports",
                offline_sample=True,
            )
        )
    )
    workbook = openpyxl.load_workbook(result.output_paths["xlsx"], read_only=True)
    headers = [cell.value for cell in next(workbook["backtest_metrics"].iter_rows(max_row=1))]

    assert {"full_total_turnover", "full_annualized_turnover", "full_total_cost", "full_annualized_cost_drag"}.issubset(headers)


def test_report_factor_score_sheets_include_eligibility_scope(tmp_path):
    result = write_reports(
        run_analysis(
            RunConfig(
                start_date="2019-01-01",
                output_dir=tmp_path / "outputs",
                report_dir=tmp_path / "reports",
                offline_sample=True,
            )
        )
    )
    workbook = openpyxl.load_workbook(result.output_paths["xlsx"], read_only=True)

    factor_headers = [cell.value for cell in next(workbook["factor_scores"].iter_rows(max_row=1))]
    history_headers = [cell.value for cell in next(workbook["factor_score_history_top20"].iter_rows(max_row=1))]
    selected_headers = [cell.value for cell in next(workbook["selected_factor_scores"].iter_rows(max_row=1))]

    assert {"eligible_for_current_model_portfolio", "score_scope"}.issubset(factor_headers)
    assert {"eligible_for_model_portfolio_at_date", "score_scope"}.issubset(history_headers)
    assert {"eligible_for_current_model_portfolio", "score_scope"}.issubset(selected_headers)


def test_liquidity_capacity_thresholds_and_counts_are_exported(tmp_path):
    result = write_reports(
        run_analysis(
            RunConfig(
                start_date="2019-01-01",
                output_dir=tmp_path / "outputs",
                report_dir=tmp_path / "reports",
                offline_sample=True,
                min_avg_volume=123,
                min_avg_dollar_volume=456,
            )
        )
    )
    output_sheet = result.metadata["recommendation_output_sheet"]
    workbook = openpyxl.load_workbook(result.output_paths["xlsx"], read_only=True)
    headers = [cell.value for cell in next(workbook[output_sheet].iter_rows(max_row=1))]

    assert result.metadata["recommendation_liquidity_status_counts"]
    assert LIQUIDITY_CAPACITY_COLUMNS.issubset(headers)
    assert WEIGHTING_COLUMNS.issubset(headers)
    assert result.recommendations["min_avg_volume_required"].eq(123.0).all()
    assert result.recommendations["min_avg_dollar_volume_required"].eq(456.0).all()
    assert result.recommendations["capacity_warning"].str.contains("Capacity is not estimated").all()



def test_research_validation_mode_records_audit_warning_not_blocker(tmp_path, monkeypatch):
    config = RunConfig(
        output_dir=tmp_path / "outputs",
        report_dir=tmp_path / "reports",
        offline_sample=False,
        stale_after_days=10_000,
        top_n=5,
        max_weight=0.2,
        factor_selection_mode="research_validation",
    )
    monkeypatch.setattr(
        "momentum_factor_lab.workflow.load_market_data",
        lambda _: _live_fixture_market_data(config, subset_run=False, point_in_time=True),
    )

    result = run_analysis(config)

    assert result.metadata["factor_selection_mode"] == "research_validation"
    assert result.metadata["selected_factor_selection_source"] == "research_validation"
    assert not result.metadata["same_sample_selection_blocked_for_tradable"]
    assert result.metadata["factor_selection_warning"]
    assert "predeclared_selected_factor" not in result.metadata["execution_limitations"]
    assert result.metadata["tradability_requirements"]["factor_selection_policy_available"]
    assert result.metadata["current_recommendations_available"]
    assert result.recommendations["weight"].gt(0.0).any()


def test_walk_forward_mode_records_prior_window_selection_history(tmp_path):
    result = run_analysis(
        RunConfig(
            start_date="2019-01-01",
            output_dir=tmp_path / "outputs",
            report_dir=tmp_path / "reports",
            offline_sample=True,
            factor_selection_mode="walk_forward",
        )
    )

    assert result.metadata["factor_selection_mode"] == "walk_forward"
    assert result.metadata["selected_factor_selection_source"] == "walk_forward"
    assert not result.metadata["same_sample_selection_blocked_for_tradable"]
    assert not result.metadata["selection_policy_frozen_for_live"]
    assert not result.selection_history.empty
    assert result.selection_history["selection_source"].eq("walk_forward").all()


def test_cost_stress_export_is_separate_from_base_metrics(tmp_path):
    result = write_reports(
        run_analysis(
            RunConfig(
                start_date="2019-01-01",
                output_dir=tmp_path / "outputs",
                report_dir=tmp_path / "reports",
                offline_sample=True,
                cost_stress_high_bps=75,
            )
        )
    )
    json_path = tmp_path / "run_results.json"
    write_run_results_json(result, json_path)
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    workbook = openpyxl.load_workbook(result.output_paths["xlsx"], read_only=True)

    assert "cost_stress" in workbook.sheetnames
    assert "cost_stress" in payload
    scenarios = {row["scenario"] for row in payload["cost_stress"]}
    assert {"base_configured_cost", "high_cost_stress"}.issubset(scenarios)
    assert all(row["stress_metric_type"] == "returns_recomputed_from_turnover" for row in payload["cost_stress"])
    assert {"stressed_sharpe", "stressed_max_drawdown", "stressed_annualized_cost_drag"}.issubset(payload["cost_stress"][0])
    assert result.metrics.loc[result.selected_factor, "full_total_cost"] == result.metadata["selected_factor_total_cost"]
