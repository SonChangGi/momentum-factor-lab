from datetime import UTC, datetime
from types import SimpleNamespace

import numpy as np
import openpyxl
import pandas as pd

from momentum_factor_lab.backtest import run_factor_backtest
from momentum_factor_lab.config import RunConfig
from momentum_factor_lab.factors import compute_factor_scores, factor_definitions_frame, validate_factor_library
from momentum_factor_lab.portfolio import balanced_weights, recommendation_table
from momentum_factor_lab.report import _factor_history_top_frame, _latest_factor_scores_frame, write_excel, write_pdf
from momentum_factor_lab.workflow import RunResult, _metrics_for_backtests, _score_factors


def test_synthetic_2000_symbol_all_factor_top20_scale_and_report(tmp_path):
    dates = pd.bdate_range("2023-01-02", periods=320)
    cols = [f"S{i:04d}" for i in range(2000)]
    rng = np.random.default_rng(7)
    style_drift = np.linspace(-0.00005, 0.00035, len(cols))[None, :]
    common = rng.normal(0.00015, 0.006, (len(dates), 1))
    idiosyncratic = rng.normal(0.0, 0.010, (len(dates), len(cols)))
    returns = style_drift + common + idiosyncratic
    prices = pd.DataFrame(50 * np.exp(np.cumsum(returns, axis=0)), index=dates, columns=cols)
    scores = compute_factor_scores(prices)
    assert len(scores) >= 55
    config = RunConfig(start_date="2023-01-02", end_date="2024-03-22", top_n=20, max_weight=0.05)
    backtests = {
        name: run_factor_backtest(prices, factor_scores, config, name)
        for name, factor_scores in scores.items()
    }
    assert all(result.weights.gt(0).sum(axis=1).max() == 20 for result in backtests.values())
    metrics, robustness = _metrics_for_backtests(backtests)
    score_components = _score_factors(metrics)
    selected = str(score_components.index[0])
    latest_scores = scores[selected].iloc[-1].dropna()
    recommendations = recommendation_table(latest_scores, balanced_weights(latest_scores, 20, 0.05), top_n=20)
    recommendations["recommendation_status"] = "synthetic_scale_test"
    recommendations["signal_date"] = str(dates[-1].date())
    benchmark_cols = [
        "factor",
        "benchmark",
        "strategy_cagr",
        "benchmark_cagr",
        "strategy_sharpe",
        "benchmark_sharpe",
        "strategy_max_drawdown",
        "benchmark_max_drawdown",
        "annualized_excess_return",
        "tracking_error",
        "information_ratio",
        "beta_to_benchmark",
    ]
    result = RunResult(
        config=config,
        market_data=SimpleNamespace(
            candidate_universe=pd.DataFrame({"symbol": cols}),
            prices=prices,
            price_sources=pd.DataFrame({"symbol": cols, "price_source": "synthetic"}),
            exclusions=pd.DataFrame(columns=["symbol", "reason", "observed"]),
        ),
        factor_scores=scores,
        backtests=backtests,
        metrics=metrics,
        score_components=score_components,
        selected_factor=selected,
        selected_reason="synthetic scale test",
        recommendations=recommendations,
        robustness=robustness,
        sensitivity=pd.DataFrame(columns=["factor", "variant_type", "variant", "parameter_set"]),
        benchmark_relative=pd.DataFrame(columns=benchmark_cols),
        factor_validation=validate_factor_library(prices),
        factor_definitions=factor_definitions_frame(),
        data_sources=pd.DataFrame([{"source": "synthetic", "status": "generated", "records": len(cols)}]),
        metadata={
            "run_timestamp_utc": datetime.now(UTC).isoformat(),
            "recommendation_status": "synthetic_scale_test",
            "provider": "synthetic",
            "data_as_of": str(dates[-1].date()),
            "candidate_universe_size": len(cols),
            "eligible_price_universe_size": len(cols),
            "excluded_symbols": 0,
            "portfolio_construction": "long-only top-20 factor portfolios",
            "survivorship_bias_caveat": "synthetic test",
            "live_data_gate": "synthetic test",
            "non_advice_disclaimer": "synthetic test",
        },
        output_paths={},
    )
    assert len(_latest_factor_scores_frame(result)) < 1_048_576
    assert len(_factor_history_top_frame(result)) < 1_048_576
    xlsx = tmp_path / "large.xlsx"
    pdf = tmp_path / "large.pdf"
    write_excel(result, xlsx)
    write_pdf(result, pdf)
    assert xlsx.exists() and pdf.exists()
    workbook = openpyxl.load_workbook(xlsx, read_only=True)
    assert workbook["factor_scores"].max_row < 1_048_576
    assert workbook["factor_score_history_top20"].max_row < 1_048_576
    assert workbook["price_sources"].max_row == 2001
